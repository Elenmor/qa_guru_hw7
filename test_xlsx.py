from openpyxl import load_workbook
from os_path_scripts import PROJECT_ROOT_PATH, resources
import os


# TODO оформить в тест, добавить ассерты и использовать универсальный путь


def test_xlsx():

    workbook = load_workbook(os.path.join(PROJECT_ROOT_PATH, resources, 'file_example_XLSX_50.xlsx'))
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)

    assert sheet.cell(row=3, column=2).value == 'Mara'

